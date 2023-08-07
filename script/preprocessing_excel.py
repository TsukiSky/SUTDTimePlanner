import openpyxl

# static variables
ORI_FILE_PATH = '2023 Trimester 1 - T4T6 (revamped).xlsx'
DIC_FILE_PATH = 'class_dictionary.xlsx'
ALL_DATA = 'T4T6'
COHORT = 'cohort'
LECTURE = 'lecture'
C_AFTER = 'c_after'
L_AFTER = 'l_after'

# load sheet
ori_workbook = openpyxl.load_workbook(ORI_FILE_PATH)
dic_workbook = openpyxl.load_workbook(DIC_FILE_PATH)
ori_sheet = ori_workbook[ALL_DATA]
c_sheet = ori_workbook[COHORT]
c_after_sheet = ori_workbook[C_AFTER]
l_after_sheet = ori_workbook[L_AFTER]
l_sheet = ori_workbook[LECTURE]
dic_sheet = dic_workbook[ALL_DATA]
ori_num_of_rows = ori_sheet.max_row
dic_num_of_row = dic_sheet.max_row

# for i in range(6, ori_num_of_rows + 1):
#     course_id = ori_sheet.cell(row=i, column=5)
#     for j in range(1, dic_num_of_row + 1):
#         course_id_dic = dic_sheet.cell(row=j, column=1)
#         if course_id.value == course_id_dic.value:
#             ori_sheet.cell(row=i,column=8).value = dic_sheet.cell(row=j, column=3).value
#             break


# ori_workbook.save('2023 Trimester 1 - T4T6 (revamped).xlsx')

# # time format
# for c in range(3, 5):
#     for i in range(6, ori_num_of_rows+1):
#         temp = ori_sheet.cell(row = i, column=c).value
#         ori_sheet.cell(row = i, column=c).value = temp+':00'

# # weekday format
# for i in range(6, ori_num_of_rows+1):
#     temp = ori_sheet.cell(row=i, column=1).value
#     ori_sheet.cell(row=i, column=1).value = temp[:3].upper()

# # add pillar
# for i in range(6, ori_num_of_rows+1):
#     if ori_sheet.cell(row=i, column=8).value[8:11] == 'asd':
#         ori_sheet.cell(row=i, column=9).value = 'ASD'
#     if ori_sheet.cell(row=i, column=8).value[8:11] == 'esd':
#         ori_sheet.cell(row=i, column=9).value = 'ESD'
#     if ori_sheet.cell(row=i, column=8).value[8:11] == 'epd':
#         ori_sheet.cell(row=i, column=9).value = 'EPD'
#     if ori_sheet.cell(row=i, column=8).value[8:11] == 'ist':
#         ori_sheet.cell(row=i, column=9).value = 'CSD'
#     if ori_sheet.cell(row=i, column=8).value[8:11] == 'dai':
#         ori_sheet.cell(row=i, column=9).value = 'DAI'        

# # distinguish cohort and lecture
# for i in range(6, ori_num_of_rows+1):
#     if ori_sheet.cell(row=i, column=7).value[0]=='C':
#         for j in range(1, 10):
#             # if j == 7:
#             #     c_sheet.cell(row=i-5, column=j).value = 'Class'
#             # elif j == 2: continue
#             # else:
#             c_sheet.cell(row=i-5, column=j).value = ori_sheet.cell(row=i, column=j).value
#     else:
#         for j in range(1, 10):
#             # if j == 7:
#             #     l_sheet.cell(row=i-5, column=j).value = 'lecture'
#             # elif j == 2: continue
#             # else:
#             l_sheet.cell(row=i-5, column=j).value = ori_sheet.cell(row=i, column=j).value

# c_index_row = []
# l_index_row = []
# for i in range(1, c_sheet.max_row+1):
#     if c_sheet.cell(row=i, column=2).value is None:
#         c_index_row.append(i)
# for row_del in range(len(c_index_row)):
#     c_sheet.delete_rows(idx=c_index_row[row_del], amount=1)
#     c_index_row = list(map(lambda k: k-1, c_index_row))

# for i in range(1, l_sheet.max_row+1):
#     if l_sheet.cell(row=i, column=2).value is None:
#         l_index_row.append(i)
# for row_del in range(len(l_index_row)):
#     l_sheet.delete_rows(idx=l_index_row[row_del], amount=1)
#     l_index_row = list(map(lambda k: k-1, l_index_row))

# # delete date and course_id columns
# # extract weekly classes
# c_rows = []
# c_results = []
# for i in range(1, c_sheet.max_row+1):
#     row = [[] for _ in range(8)]
#     for j in range(1, 8):
#         row[j] = c_sheet.cell(row=i, column=j).value
#     c_rows.append(row)
# for row in c_rows:
#     if c_rows.count(row)>3 and row not in c_results:
#         c_results.append(row)
# for row in range(len(c_results)):
#     for col in range(1, len(c_results[row])):
#         temp = c_results[row]
#         ans = temp[col]
#         c_after_sheet.cell(row=row+1, column=col).value = ans

l_rows = []
l_results = []
for i in range(1, l_sheet.max_row+1):
    row = [[] for _ in range(8)]
    for j in range(1, 8):
        row[j] = l_sheet.cell(row=i, column=j).value
    l_rows.append(row)
for row in l_rows:
    if l_rows.count(row)>3 and row not in l_results:
        l_results.append(row)
for row in range(len(l_results)):
    for col in range(1, len(l_results[row])):
        temp = l_results[row]
        ans = temp[col]
        l_after_sheet.cell(row=row+1, column=col).value = ans

ori_workbook.save('2023 Trimester 1 - T4T6 (revamped).xlsx')
