import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import time

# static variables
FILE_PATH = '/home/zhuoran/Downloads/Telegram Desktop/2023 Trimester 1- T4T6 (2) (copy).xlsx'
WAITING_TIME = 3
SHEET = 'T4T6'

# set up chrome driver
# ser = Service("chromedriver")
ser = Service("/home/zhuoran/Downloads/chromedriver_linux64/chromedriver")
chrome_options = Options()
chrome_options.add_argument("--headless=True")
chrome_options.add_argument("--incognito")
chrome_options.add_argument("--window-size=1920x1080")
driver = webdriver.Chrome(service=ser, options=chrome_options)  # initiate web driver

# load sheet
workbook = openpyxl.load_workbook(FILE_PATH)
sheet = workbook[SHEET]
num_of_rows = sheet.max_row

for i in range(6, num_of_rows + 1):
    course_id = str(sheet.cell(row=i, column=5).value)
    if len(course_id) < 6:
        course_id = course_id + '0'  # fill up ending 0s
    url = f'https://sutd.edu.sg/search?q={course_id}'
    driver.get(url)
    time.sleep(WAITING_TIME)

    driver.find_element(By.XPATH, '/html/body/form/section/div/article/div/div/div/div/div[5]/div[2]/div/div/div[1]/div[1]/div[1]/div[1]/div/a').click()
    driver.switch_to.window(driver.window_handles[1])   # switch to new tab

    current_url = driver.current_url    # get course link

    # get course name
    if current_url[0:3] == 'esd':
        # ESD
        course_name = driver.find_element(By.CLASS_NAME, 'fusion-post-title').text
    else:
        # ISTD / DAI / EPD / ASD
        course_name = driver.find_element(By.CLASS_NAME, 'entry-title').text

    driver.close()
    driver.switch_to.window(driver.window_handles[0])
    time.sleep(WAITING_TIME)

    sheet.cell(row=i, column=6).value = course_name
    sheet.cell(row=i, column=7).value = current_url

workbook.save('2023 Trimester 1 - T4T6 (revamped).xlsx')

# for c in range(3, 5):
#     for i in range(6, m_row+1):
#         temp = sheet_obj.cell(row = i, column=c).value
#         sheet_obj.cell(row = i, column=c).value = temp+':00'

# for i in range(6, m_row+1):
#     temp = sheet_obj.cell(row=i, column=1).value
#     sheet_obj.cell(row=i, column=1).value = temp[:3].upper()
