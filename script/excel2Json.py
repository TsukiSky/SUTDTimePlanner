import json
import openpyxl

# static variables
ORI_FILE_PATH = '2023 Trimester 1 - T4T6 (revamped).xlsx'
SHEET = 'c_after'
L_SHEET = 'l_after'

ori_workbook = openpyxl.load_workbook(ORI_FILE_PATH)
sheet = ori_workbook[SHEET]
l_sheet = ori_workbook[L_SHEET]
num_of_rows = sheet.max_row
l_num_rows = l_sheet.max_row

# full_list = []
# for i in range(1, num_of_rows+1):
#     excel_row = []
#     for j in range(1, 8):
#         excel_row.append(sheet.cell(row=i, column=j).value)
#     full_list.append(excel_row)

def lec_or_cohort(section):
    if section[0] == 'L':
        return 'lecture'
    else:
        return 'Class'

# result = []

# # Create the nested JSON
# pre = full_list[0] #cur row
# pre_json = {
#         "name": full_list[0][3],
#         "pillar": full_list[0][6],
#         "link": full_list[0][5],
#         "classes": [
#             {
#                 "slots": [
#                     {
#                         "type": lec_or_cohort(full_list[0][4]),
#                         "date": full_list[0][0],
#                         "startTime": full_list[0][1],
#                         "endTime": full_list[0][2]
#                     }
#                 ]
#             }
#         ],
#         "terms": [
#             {
#                 "term": 6,
#                 "period": 'summer'
#             }
#         ]
#     }
# result.append(pre_json)
# # Convert the Python dictionary to JSON
# json_string = json.dumps(pre_json, indent=2)
# # print(json_string)

# # Save JSON to a file
# with open('output.json', 'a') as json_file:
#     json.dump(pre_json, json_file, indent=2)
#     json_file.write(',')

# for i in range(1, num_of_rows):
#     if full_list[i][3] == pre[3]: #same course
#         print(pre[3])
#         if full_list[i][4] == pre[4]:  #same section
#             print(pre[4])
#             pre_json["classes"][0]['slots'].append(
#             {
#                 "type": lec_or_cohort(full_list[i][4]),
#                 "date": full_list[i][0],
#                 "startTime": full_list[i][1],
#                 "endTime": full_list[i][2]
#             }
#             )
#             result[0] = pre_json
#             print(result[0])
#         else:
#             pre_json['classes'].append(
#             {
#                 "slots": [
#                     {
#                         "type": lec_or_cohort(full_list[i][4]),
#                         "date": full_list[i][0],
#                         "startTime": full_list[i][1],
#                         "endTime": full_list[i][2]
#                     }
#                 ]
#             } 
#             )
#             result[0] = pre_json
#             pre[4] = full_list[i][4]
#             print(pre[4])
#     else:
#         json_data = {
#             "name": full_list[i][3],
#             "pillar": full_list[i][6],
#             "link": full_list[i][5],
#             "classes": [
#                 {
#                     "slots": [
#                         {
#                             "type": lec_or_cohort(full_list[i][4]),
#                             "date": full_list[i][0],
#                             "startTime": full_list[i][1],
#                             "endTime": full_list[i][2]
#                         }
#                     ]
#                 }
#             ],
#             "terms": [
#                 {
#                     "term": 6,
#                     "period": 'summer'
#                 }
#             ]
#         }
#         pre[3] = full_list[i][3]
#         pre[4] = full_list[i][4]
#         result.insert(0, json_data)
#         pre_json = json_data


# for data in result:
#     # Convert the Python dictionary to JSON
#     json_string = json.dumps(data, indent=2)
#     # print(json_string)

#     # Save JSON to a file
#     with open('output.json', 'a') as json_file:
#         json.dump(data, json_file, indent=2)
#         if i!=len(result)-1:
#             json_file.write(',')

# # delete first item, add [], delete last ','

with open('output.json', 'r') as file:
    # read JSON data
    data = json.load(file)

    for item in data:
        for i in range(1, l_num_rows+1):
            if item['name'] == l_sheet.cell(row=i, column=4).value:
                for slot in item['classes']:
                    slot['slots'].append(
                    {
                        "type": lec_or_cohort(l_sheet.cell(row=i, column=5).value),
                        "date": l_sheet.cell(row=i, column=1).value,
                        "startTime": l_sheet.cell(row=i, column=2).value,
                        "endTime": l_sheet.cell(row=i, column=3).value
                    }
                    ) 

    newData = json.dumps(data, indent=2)

with open('modified.json', 'w') as file:
    # write
    file.write(newData)