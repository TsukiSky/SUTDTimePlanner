import openpyxl

# static variables
LINK_FILE_PATH = '/home/zhuoran/Downloads/Telegram Desktop/2023 Trimester 1- T4T6 (2) (copy).xlsx'
DIC_FILE_PATH = '/home/zhuoran/Downloads/Telegram Desktop/class_dictionary.xlsx'
SHEET = 'T4T6'

# load sheet
link_workbook = openpyxl.load_workbook(LINK_FILE_PATH)
dic_workbook = openpyxl.load_workbook(DIC_FILE_PATH)
link_sheet = link_workbook[SHEET]
dic_sheet = dic_workbook[SHEET]
link_num_of_rows = link_sheet.max_row
dic_num_of_rows = dic_sheet.max_row

for i in range(1, dic_num_of_rows + 1):
    course_id = str(dic_sheet.cell(row=i, column=1).value)
    for j in range(6, link_num_of_rows+1):
        course_id_link = str(link_sheet.cell(row=j, column=5).value)
        if course_id == course_id_link:
            dic_sheet.cell(row=i, column=3).value = link_sheet.cell(row=j, column=8).value
            break

dic_workbook.save('2023 Trimester 1 - T4T6 (revamped).xlsx')

# for c in range(3, 5):
#     for i in range(6, m_row+1):
#         temp = sheet_obj.cell(row = i, column=c).value
#         sheet_obj.cell(row = i, column=c).value = temp+':00'

# for i in range(6, m_row+1):
#     temp = sheet_obj.cell(row=i, column=1).value
#     sheet_obj.cell(row=i, column=1).value = temp[:3].upper()
